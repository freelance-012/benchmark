"""Rebuild one simple run-level XLSX summary from saved Segment facts."""

from __future__ import annotations

import os
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

import yaml

from ..algorithms.contracts import (
    EVALUATION_WORKFLOW_SF_VLOC,
    EVALUATION_WORKFLOW_SF_VO,
    EvaluationWorkflowConfig,
)
from ..debug import debug_input, debug_output
from .models import (
    DEFAULT_RPE_DELTA_UNIT,
    DEFAULT_RPE_DELTA_VALUE,
    normalize_rpe_delta,
)
from .service import (
    EvaluationError,
    VLOC_METRIC_KEYS,
    VO_METRIC_KEYS,
    load_metrics,
    metric_values,
)

SUMMARY_FILENAME = "run_summary.xlsx"


def _generate_timestamped_filename() -> str:
    """Generate a timestamped filename for the summary report."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"run_summary_{timestamp}.xlsx"


@dataclass(frozen=True)
class _SummaryRow:
    run_index: int
    result_path: Path
    dataset_path: Path
    run_status: str
    evaluation_status: str
    failure_reason: Optional[str]
    metrics: Tuple[Optional[float], ...]


class SummaryWorkbookWriter:
    def generate_from_test(
        self, test_root: Path, *, use_timestamp: bool = False
    ) -> Path:
        root = test_root.expanduser().resolve()
        if not root.is_dir():
            raise EvaluationError(f"test directory does not exist: {root}")
        algorithm_config = _load_yaml(root / "config" / "algorithm.yaml")
        workflows = _workflows_from_algorithm_config(algorithm_config)
        return self.update(root, workflows, use_timestamp=use_timestamp)

    def update(
        self,
        test_root: Path,
        workflows: Sequence[EvaluationWorkflowConfig],
        *,
        use_timestamp: bool = False,
    ) -> Path:
        for wf in workflows:
            if wf.workflow not in {
                EVALUATION_WORKFLOW_SF_VO,
                EVALUATION_WORKFLOW_SF_VLOC,
            }:
                raise EvaluationError(f"unsupported summary workflow: {wf.workflow}")
        root = test_root.expanduser().resolve()
        debug_input(
            "REPORT",
            source=root / "dataset",
            workflows=[w.directory_name for w in workflows],
        )
        rpe_delta_value, rpe_delta_unit = _rpe_delta_from_run_config(
            _load_yaml(root / "config" / "run.yaml")
        )
        sheet_data: List[Tuple[str, str, Tuple[_SummaryRow, ...]]] = []
        for wf in workflows:
            rows = self._load_rows(root, wf)
            sheet_data.append((wf.directory_name, wf.workflow, rows))
        output_filename = (
            _generate_timestamped_filename() if use_timestamp else SUMMARY_FILENAME
        )
        output_path = root / output_filename
        self._write_atomic(
            output_path,
            sheet_data,
            rpe_delta_value,
            rpe_delta_unit,
        )
        debug_output(
            "REPORT",
            input=root / "dataset",
            sheets=len(sheet_data),
            saved=output_path,
        )
        return output_path

    def _load_rows(
        self,
        test_root: Path,
        workflow_config: EvaluationWorkflowConfig,
    ) -> Tuple[_SummaryRow, ...]:
        run_config = _load_yaml(test_root / "config" / "run.yaml")
        rpe_delta_value, rpe_delta_unit = _rpe_delta_from_run_config(run_config)
        dataset_paths = _dataset_paths_from_run_config(run_config)
        raw_order = run_config.get("segment_order")
        if not isinstance(raw_order, list):
            raise EvaluationError("frozen run configuration has no segment_order")

        workflow = workflow_config.workflow
        rows: List[_SummaryRow] = []
        for item in raw_order:
            if not isinstance(item, dict) or "run_index" not in item:
                raise EvaluationError("frozen segment_order contains an invalid item")
            try:
                run_index = int(item["run_index"])
            except (TypeError, ValueError) as exc:
                raise EvaluationError(f"frozen run_index is invalid") from exc
            dataset_id = item.get("dataset_id")
            if not isinstance(dataset_id, str) or dataset_id not in dataset_paths:
                raise EvaluationError(
                    "frozen segment_order references an unknown dataset"
                )
            segment_dir = (test_root / "dataset" / str(run_index)).resolve()
            run_status = "not_run"
            evaluation_status = "not_run"
            failure_reason: Optional[str] = None
            values: Tuple[Optional[float], ...] = tuple(
                None for _ in _metric_keys(workflow)
            )

            run_receipt_path = segment_dir / "receipt.yaml"
            result_path = segment_dir
            if run_receipt_path.is_file():
                try:
                    run_receipt = _load_yaml(run_receipt_path)
                except EvaluationError as exc:
                    run_status = "failed"
                    failure_reason = str(exc)
                else:
                    run_status = str(run_receipt.get("status", "failed"))
                    raw_reason = run_receipt.get("failure_reason")
                    if raw_reason is not None:
                        failure_reason = str(raw_reason)
                    output_source_paths = run_receipt.get("output_source_paths", [])
                    if output_source_paths:
                        result_path = Path(output_source_paths[0]).parent

            eval_subdir = segment_dir / "evaluation" / workflow_config.directory_name
            evaluation_receipt_path = eval_subdir / "receipt.yaml"
            if run_status == "success" and evaluation_receipt_path.is_file():
                try:
                    evaluation_receipt = _load_yaml(evaluation_receipt_path)
                except EvaluationError as exc:
                    evaluation_status = "failed"
                    failure_reason = str(exc)
                else:
                    evaluation_status = str(
                        evaluation_receipt.get("status", "failed")
                    )
                    raw_reason = evaluation_receipt.get("failure_reason")
                    if raw_reason is not None:
                        failure_reason = str(raw_reason)

                    if evaluation_status == "success":
                        metrics_path = eval_subdir / "metrics.json"
                        try:
                            metrics, _ = load_metrics(
                                metrics_path,
                                workflow,
                                expected_delta_value=rpe_delta_value,
                                expected_delta_unit=rpe_delta_unit,
                            )
                            values = metric_values(metrics, workflow)
                        except EvaluationError as exc:
                            evaluation_status = "failed"
                            failure_reason = str(exc)

            rows.append(
                _SummaryRow(
                    run_index=run_index,
                    result_path=result_path,
                    dataset_path=dataset_paths[dataset_id],
                    run_status=run_status,
                    evaluation_status=evaluation_status,
                    failure_reason=failure_reason,
                    metrics=values,
                )
            )
        return tuple(sorted(rows, key=lambda row: row.run_index))

    @staticmethod
    def _write_atomic(
        output_path: Path,
        sheet_data: Sequence[Tuple[str, str, Tuple[_SummaryRow, ...]]],
        rpe_delta_value: float,
        rpe_delta_unit: str,
    ) -> None:
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise EvaluationError(
                "openpyxl is required to generate run_summary.xlsx"
            ) from exc

        workbook = Workbook()
        workbook.remove(workbook.active)

        for sheet_name, workflow, rows in sheet_data:
            _write_sheet(
                workbook,
                sheet_name,
                workflow,
                rows,
                rpe_delta_value,
                rpe_delta_unit,
            )

        temporary: Optional[Path] = None
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            with tempfile.NamedTemporaryFile(
                dir=output_path.parent,
                prefix=f".{output_path.name}.",
                suffix=".tmp.xlsx",
                delete=False,
            ) as handle:
                temporary = Path(handle.name)
            workbook.save(temporary)
            os.replace(temporary, output_path)
        except OSError as exc:
            raise EvaluationError(
                f"cannot save summary workbook {output_path}: {exc}"
            ) from exc
        finally:
            workbook.close()
            if temporary is not None and temporary.exists():
                temporary.unlink(missing_ok=True)


def _write_sheet(
    workbook: Any,
    sheet_name: str,
    workflow: str,
    rows: Tuple[_SummaryRow, ...],
    rpe_delta_value: float,
    rpe_delta_unit: str,
) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    sheet = workbook.create_sheet(title=sheet_name)
    sheet.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    red_fill = PatternFill("solid", fgColor="FFC7CE")

    metric_keys = _metric_keys(workflow)
    metric_start_column = 4
    if workflow == EVALUATION_WORKFLOW_SF_VO:
        metric_titles = ("RMSE", "Mean", "Median", "Max", "Min", "Count")
        sheet.merge_cells("A1:A2")
        sheet.merge_cells("B1:B2")
        sheet.merge_cells("C1:C2")
        sheet.merge_cells("D1:I1")
        sheet.merge_cells("J1:J2")
        sheet["A1"] = "运行编号"
        sheet["B1"] = "结果路径"
        sheet["C1"] = "数据集路径"
        sheet["D1"] = (
            "RPE 平移误差 "
            f"(delta={_format_rpe_delta(rpe_delta_value, rpe_delta_unit)})"
        )
        sheet["J1"] = "Segment 数量"
        for column, title in enumerate(
            metric_titles,
            start=metric_start_column,
        ):
            sheet.cell(row=2, column=column, value=title)
        status_column = 11
        data_start_row = 3
        for column, title in enumerate(
            ("运行状态", "评估状态", "失败原因"),
            start=status_column,
        ):
            start = sheet.cell(row=1, column=column, value=title)
            sheet.merge_cells(
                start_row=1,
                start_column=column,
                end_row=2,
                end_column=column,
            )
            start.alignment = center
        sheet.freeze_panes = "A3"
    else:
        headers = (
            "运行编号",
            "结果路径",
            "数据集路径",
            *metric_keys,
            "运行状态",
            "评估状态",
            "失败原因",
        )
        for column, title in enumerate(headers, start=1):
            sheet.cell(row=1, column=column, value=title)
        status_column = metric_start_column + len(metric_keys)
        data_start_row = 2
        sheet.freeze_panes = "A2"

    max_header_row = data_start_row - 1
    total_columns = status_column + 2
    for row in sheet.iter_rows(
        min_row=1,
        max_row=max_header_row,
        min_col=1,
        max_col=total_columns,
    ):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

    for row_offset, summary in enumerate(rows):
        row_number = data_start_row + row_offset
        sheet.cell(row=row_number, column=1, value=summary.run_index)
        path_cell = sheet.cell(
            row=row_number,
            column=2,
            value=str(summary.result_path),
        )
        path_cell.hyperlink = summary.result_path.as_uri()
        path_cell.style = "Hyperlink"
        dataset_cell = sheet.cell(
            row=row_number,
            column=3,
            value=str(summary.dataset_path),
        )
        dataset_cell.hyperlink = summary.dataset_path.as_uri()
        dataset_cell.style = "Hyperlink"

        for metric_offset, value in enumerate(
            summary.metrics,
            start=metric_start_column,
        ):
            cell = sheet.cell(row=row_number, column=metric_offset, value=value)
            if value is not None:
                metric_key = metric_keys[metric_offset - metric_start_column]
                if metric_key in {"count", "segment_count"}:
                    cell.number_format = "0"
                else:
                    cell.number_format = "0.000000"

        sheet.cell(
            row=row_number,
            column=status_column,
            value=summary.run_status,
        )
        sheet.cell(
            row=row_number,
            column=status_column + 1,
            value=summary.evaluation_status,
        )
        sheet.cell(
            row=row_number,
            column=status_column + 2,
            value=summary.failure_reason,
        )

        if workflow == EVALUATION_WORKFLOW_SF_VLOC:
            horizontal_error = sheet.cell(
                row=row_number,
                column=metric_start_column
                + metric_keys.index("mean_error_pos_xy"),
            )
            if horizontal_error.value is not None:
                if float(horizontal_error.value) > 50.0:
                    horizontal_error.fill = red_fill
                elif float(horizontal_error.value) > 20.0:
                    horizontal_error.fill = yellow_fill

        for column in range(1, total_columns + 1):
            cell = sheet.cell(row=row_number, column=column)
            cell.border = border
            cell.alignment = left if column in {2, 3, total_columns} else center

    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 58
    sheet.column_dimensions["C"].width = 58
    for column in range(metric_start_column, status_column):
        sheet.column_dimensions[_column_letter(column)].width = 22
    sheet.column_dimensions[_column_letter(status_column)].width = 14
    sheet.column_dimensions[_column_letter(status_column + 1)].width = 14
    sheet.column_dimensions[_column_letter(status_column + 2)].width = 48
    sheet.row_dimensions[1].height = 24
    if workflow == EVALUATION_WORKFLOW_SF_VO:
        sheet.row_dimensions[2].height = 22
    if rows and workflow == EVALUATION_WORKFLOW_SF_VLOC:
        sheet.auto_filter.ref = (
            f"A1:{_column_letter(total_columns)}{data_start_row + len(rows) - 1}"
        )


def _load_yaml(path: Path) -> Dict[str, Any]:
    try:
        payload: Any = yaml.safe_load(path.read_text(encoding="utf-8"))
    except (OSError, yaml.YAMLError, ValueError) as exc:
        raise EvaluationError(f"cannot read YAML {path}: {exc}") from exc
    if not isinstance(payload, dict):
        raise EvaluationError(f"YAML root must be a mapping: {path}")
    return payload


def _metric_keys(workflow: str) -> Tuple[str, ...]:
    if workflow == EVALUATION_WORKFLOW_SF_VO:
        return VO_METRIC_KEYS
    if workflow == EVALUATION_WORKFLOW_SF_VLOC:
        return VLOC_METRIC_KEYS
    raise EvaluationError(f"unsupported summary workflow: {workflow}")


def _workflows_from_algorithm_config(
    config: Mapping[str, Any],
) -> Tuple[EvaluationWorkflowConfig, ...]:
    contract = config.get("contract")
    if not isinstance(contract, Mapping):
        raise EvaluationError(
            "frozen algorithm configuration has no contract mapping"
        )
    raw_workflows = contract.get("evaluation_workflows")
    if isinstance(raw_workflows, list) and raw_workflows:
        workflows = []
        for item in raw_workflows:
            if not isinstance(item, Mapping):
                raise EvaluationError("evaluation_workflows contains an invalid item")
            workflow = item.get("workflow")
            if not isinstance(workflow, str) or workflow not in {
                EVALUATION_WORKFLOW_SF_VO,
                EVALUATION_WORKFLOW_SF_VLOC,
            }:
                raise EvaluationError(
                    f"evaluation_workflows has unsupported workflow: {workflow!r}"
                )
            vo_filename = item.get("vo_filename")
            if vo_filename is not None and not isinstance(vo_filename, str):
                raise EvaluationError("vo_filename must be a string or null")
            workflows.append(
                EvaluationWorkflowConfig(workflow=workflow, vo_filename=vo_filename)
            )
        return tuple(workflows)
    workflow = contract.get("evaluation_workflow")
    if not isinstance(workflow, str) or workflow not in {
        EVALUATION_WORKFLOW_SF_VO,
        EVALUATION_WORKFLOW_SF_VLOC,
    }:
        raise EvaluationError(
            "frozen algorithm contract has no supported evaluation workflow"
        )
    return (EvaluationWorkflowConfig(workflow=workflow, vo_filename=None),)


def _rpe_delta_from_run_config(config: Mapping[str, Any]) -> Tuple[float, str]:
    raw_evaluation = config.get("evaluation")
    if raw_evaluation is None:
        return DEFAULT_RPE_DELTA_VALUE, DEFAULT_RPE_DELTA_UNIT
    if not isinstance(raw_evaluation, Mapping):
        raise EvaluationError("frozen run evaluation configuration is invalid")
    try:
        return normalize_rpe_delta(
            raw_evaluation["rpe_delta_value"],
            raw_evaluation["rpe_delta_unit"],
        )
    except (KeyError, ValueError) as exc:
        raise EvaluationError(
            f"frozen run RPE delta configuration is invalid: {exc}"
        ) from exc


def _dataset_paths_from_run_config(
    config: Mapping[str, Any],
) -> Dict[str, Path]:
    raw_datasets = config.get("dataset_order")
    if not isinstance(raw_datasets, list):
        raise EvaluationError("frozen run configuration has no dataset_order")

    dataset_paths: Dict[str, Path] = {}
    for item in raw_datasets:
        if not isinstance(item, Mapping):
            raise EvaluationError("frozen dataset_order contains an invalid item")
        dataset_id = item.get("dataset_id")
        root_path = item.get("root_path")
        if not isinstance(dataset_id, str) or not isinstance(root_path, str):
            raise EvaluationError("frozen dataset_order contains an invalid item")
        dataset_path = Path(root_path).expanduser()
        if not dataset_path.is_absolute():
            raise EvaluationError("frozen dataset path must be absolute")
        if dataset_id in dataset_paths:
            raise EvaluationError("frozen dataset_order contains a duplicate dataset")
        dataset_paths[dataset_id] = dataset_path.resolve()
    return dataset_paths


def _format_rpe_delta(value: float, unit: str) -> str:
    if unit == "m" and value.is_integer():
        number = f"{value:.1f}"
    else:
        number = f"{value:.15g}"
    return f"{number}{unit}"


def _column_letter(column: int) -> str:
    value = column
    letters = ""
    while value:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters
