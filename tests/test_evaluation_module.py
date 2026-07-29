from __future__ import annotations

import io
import sys
import tempfile
import unittest
from contextlib import redirect_stderr, redirect_stdout
from pathlib import Path
from typing import Optional

import yaml
from openpyxl import load_workbook

from slam_benchmark.cli import main
from slam_benchmark.evaluation import (
    EvaluationError,
    EvaluationRequest,
    EvaluationService,
    SummaryWorkbookWriter,
    normalize_rpe_delta,
)


class EvaluationModuleTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.addCleanup(self.temporary.cleanup)
        self.root = Path(self.temporary.name).resolve()
        self.test_root = self.root / "result" / "algorithm" / "test-000"
        (self.test_root / "config").mkdir(parents=True)
        self.fake_voeval = self._write_fake_voeval()
        self.service = EvaluationService(
            (sys.executable, str(self.fake_voeval)),
            timeout_seconds=10,
        )
        self.writer = SummaryWorkbookWriter()

    def test_sf_vo_writes_evaluation_facts_and_two_level_summary(self) -> None:
        self._freeze_segment_order(1)
        segment_dir = self._write_run_receipt(0, "success")
        receipt = self.service.evaluate(
            self._request(0, "sf_vo", segment_dir)
        )

        workbook_path = self.writer.update(self.test_root, "sf_vo")

        evaluation_dir = segment_dir / "evaluation"
        self.assertEqual(receipt.status, "success")
        self.assertTrue((evaluation_dir / "metrics.json").is_file())
        self.assertTrue((evaluation_dir / "receipt.yaml").is_file())
        self.assertIn(
            "fake voeval sf_vo",
            (evaluation_dir / "voeval.log").read_text(encoding="utf-8"),
        )

        workbook = load_workbook(workbook_path)
        self.addCleanup(workbook.close)
        sheet = workbook["Summary"]
        self.assertEqual(sheet["A1"].value, "运行编号")
        self.assertEqual(sheet["B1"].value, "结果路径")
        self.assertEqual(sheet["C1"].value, "数据集路径")
        self.assertEqual(sheet["D1"].value, "RPE 平移误差 (delta=100.0m)")
        self.assertEqual(
            [sheet.cell(row=2, column=column).value for column in range(4, 10)],
            ["RMSE", "Mean", "Median", "Max", "Min", "Count"],
        )
        self.assertEqual(sheet["A3"].value, 0)
        self.assertEqual(sheet["B3"].value, str(segment_dir))
        self.assertTrue(sheet["B3"].hyperlink.target.startswith("file:///"))
        self.assertEqual(sheet["C3"].value, str(self.root / "data-0"))
        self.assertTrue(sheet["C3"].hyperlink.target.startswith("file:///"))
        self.assertEqual(
            [sheet.cell(row=3, column=column).value for column in range(4, 10)],
            [1, 2, 3, 4, 0.5, 6],
        )
        self.assertEqual(sheet["J1"].value, "Segment 数量")
        self.assertEqual(sheet["J3"].value, 3)
        self.assertEqual(sheet["K3"].value, "success")
        self.assertEqual(sheet["L3"].value, "success")
        self.assertIsNone(sheet["M3"].value)

    def test_report_cli_rebuilds_excel_from_saved_test_data(self) -> None:
        self._freeze_algorithm_contract("sf_vo")
        self._freeze_segment_order(1)
        segment_dir = self._write_run_receipt(0, "success")
        self.service.evaluate(self._request(0, "sf_vo", segment_dir))
        self.fake_voeval.unlink()

        output = io.StringIO()
        errors = io.StringIO()
        with redirect_stdout(output), redirect_stderr(errors):
            exit_code = main(
                [
                    "report",
                    "--test-dir",
                    str(self.test_root),
                ]
            )

        workbook_path = self.test_root / "run_summary.xlsx"
        self.assertEqual(exit_code, 0)
        self.assertTrue(workbook_path.is_file())
        self.assertIn(str(workbook_path), output.getvalue())
        workbook = load_workbook(workbook_path)
        self.addCleanup(workbook.close)
        sheet = workbook["Summary"]
        self.assertEqual(sheet["C3"].value, str(self.root / "data-0"))
        self.assertEqual(sheet["D3"].value, 1)

    def test_report_cli_rejects_missing_test_directory(self) -> None:
        output = io.StringIO()
        errors = io.StringIO()
        missing = self.root / "missing-test"

        with redirect_stdout(output), redirect_stderr(errors):
            exit_code = main(["report", "--test-dir", str(missing)])

        self.assertEqual(exit_code, 2)
        self.assertIn("test directory does not exist", errors.getvalue())
        self.assertFalse((missing / "run_summary.xlsx").exists())

    def test_custom_frame_delta_reaches_voeval_metrics_and_workbook(self) -> None:
        self._freeze_segment_order(1, rpe_delta_value=5, rpe_delta_unit="f")
        segment_dir = self._write_run_receipt(0, "success")
        receipt = self.service.evaluate(
            self._request(
                0,
                "sf_vo",
                segment_dir,
                rpe_delta_value=5,
                rpe_delta_unit="f",
            )
        )

        workbook_path = self.writer.update(self.test_root, "sf_vo")
        delta_index = receipt.command.index("--delta")
        unit_index = receipt.command.index("--unit")
        self.assertEqual(receipt.command[delta_index + 1], "5")
        self.assertEqual(receipt.command[unit_index + 1], "f")
        metrics = self._yaml(segment_dir / "evaluation" / "metrics.json")
        self.assertEqual(metrics["rpe_translation_m"]["delta_value"], 5)
        self.assertEqual(metrics["rpe_translation_m"]["delta_unit"], "f")

        workbook = load_workbook(workbook_path)
        self.addCleanup(workbook.close)
        self.assertEqual(
            workbook["Summary"]["D1"].value,
            "RPE 平移误差 (delta=5f)",
        )

    def test_rpe_delta_validation_rejects_invalid_values(self) -> None:
        self.assertEqual(normalize_rpe_delta(25.5, "m"), (25.5, "m"))
        self.assertEqual(normalize_rpe_delta(5, "f"), (5.0, "f"))
        with self.assertRaisesRegex(ValueError, "positive finite"):
            normalize_rpe_delta(0, "m")
        with self.assertRaisesRegex(ValueError, "positive integer"):
            normalize_rpe_delta(1.5, "f")

        self._freeze_segment_order(1)
        segment_dir = self._write_run_receipt(0, "success")
        with self.assertRaisesRegex(EvaluationError, "positive finite"):
            self.service.evaluate(
                self._request(
                    0,
                    "sf_vo",
                    segment_dir,
                    rpe_delta_value=0,
                )
            )

    def test_sf_vloc_colors_horizontal_error_and_keeps_failed_run(self) -> None:
        self._freeze_segment_order(6)
        for run_index in range(5):
            segment_dir = self._write_run_receipt(run_index, "success")
            receipt = self.service.evaluate(
                self._request(run_index, "sf_vloc", segment_dir)
            )
            self.assertEqual(receipt.status, "success")
        failed_dir = self._write_run_receipt(
            5,
            "failed",
            failure_reason="algorithm failed",
        )

        workbook_path = self.writer.update(self.test_root, "sf_vloc")

        workbook = load_workbook(workbook_path)
        self.addCleanup(workbook.close)
        sheet = workbook["Summary"]
        self.assertEqual(
            [sheet.cell(row=1, column=column).value for column in range(1, 11)],
            [
                "运行编号",
                "结果路径",
                "数据集路径",
                "trajectory_length_m",
                "mean_error_pos_xy",
                "mean_error_pos_z",
                "mean_error_euler",
                "max_error_pos_xy",
                "max_error_pos_z",
                "max_error_euler",
            ],
        )
        self.assertEqual(sheet["E2"].value, 10)
        self.assertEqual(sheet["E3"].value, 20)
        self.assertEqual(sheet["E4"].value, 25)
        self.assertEqual(sheet["E5"].value, 50)
        self.assertEqual(sheet["E6"].value, 55)
        self.assertNotEqual(sheet["E2"].fill.fill_type, "solid")
        self.assertNotEqual(sheet["E3"].fill.fill_type, "solid")
        self.assertTrue(sheet["E4"].fill.fgColor.rgb.endswith("FFF2CC"))
        self.assertTrue(sheet["E5"].fill.fgColor.rgb.endswith("FFF2CC"))
        self.assertTrue(sheet["E6"].fill.fgColor.rgb.endswith("FFC7CE"))
        self.assertEqual(sheet["K7"].value, "failed")
        self.assertEqual(sheet["L7"].value, "not_run")
        self.assertEqual(sheet["M7"].value, "algorithm failed")
        self.assertEqual(sheet["B7"].value, str(failed_dir))
        self.assertEqual(sheet["C7"].value, str(self.root / "data-5"))

    def test_voeval_failure_is_recorded_without_metrics(self) -> None:
        self._freeze_segment_order(1)
        segment_dir = self._write_run_receipt(0, "success")
        data_dir = self.root / "force-evaluation-failure"
        data_dir.mkdir()

        receipt = self.service.evaluate(
            self._request(0, "sf_vo", segment_dir, data_dir=data_dir)
        )
        workbook_path = self.writer.update(self.test_root, "sf_vo")

        evaluation_dir = segment_dir / "evaluation"
        self.assertEqual(receipt.status, "failed")
        self.assertEqual(receipt.exit_code, 7)
        self.assertFalse((evaluation_dir / "metrics.json").exists())
        self.assertIn(
            "intentional evaluator failure",
            (evaluation_dir / "voeval.log").read_text(encoding="utf-8"),
        )
        saved_receipt = self._yaml(evaluation_dir / "receipt.yaml")
        self.assertEqual(saved_receipt["status"], "failed")
        self.assertIn("code 7", saved_receipt["failure_reason"])

        workbook = load_workbook(workbook_path)
        self.addCleanup(workbook.close)
        sheet = workbook["Summary"]
        self.assertEqual(sheet["K3"].value, "success")
        self.assertEqual(sheet["L3"].value, "failed")
        self.assertIn("code 7", sheet["M3"].value)
        self.assertIsNone(sheet["D3"].value)

    def test_summary_lists_planned_but_not_run_segment(self) -> None:
        self._freeze_segment_order(2)
        self._write_run_receipt(0, "failed", failure_reason="first run failed")

        workbook_path = self.writer.update(self.test_root, "sf_vo")

        workbook = load_workbook(workbook_path)
        self.addCleanup(workbook.close)
        sheet = workbook["Summary"]
        self.assertEqual(sheet["A3"].value, 0)
        self.assertEqual(sheet["K3"].value, "failed")
        self.assertEqual(sheet["L3"].value, "not_run")
        self.assertEqual(sheet["A4"].value, 1)
        self.assertEqual(sheet["K4"].value, "not_run")
        self.assertEqual(sheet["L4"].value, "not_run")

    def _freeze_segment_order(
        self,
        count: int,
        *,
        rpe_delta_value: float = 100.0,
        rpe_delta_unit: str = "m",
    ) -> None:
        (self.test_root / "config" / "run.yaml").write_text(
            yaml.safe_dump(
                {
                    "evaluation": {
                        "rpe_delta_value": rpe_delta_value,
                        "rpe_delta_unit": rpe_delta_unit,
                    },
                    "dataset_order": [
                        {
                            "dataset_id": f"dataset-{run_index}",
                            "dataset_type": "rk3399",
                            "root_path": str(self.root / f"data-{run_index}"),
                        }
                        for run_index in range(count)
                    ],
                    "segment_order": [
                        {
                            "run_index": run_index,
                            "dataset_id": f"dataset-{run_index}",
                            "segment_id": f"segment-{run_index}",
                        }
                        for run_index in range(count)
                    ],
                },
                sort_keys=False,
            ),
            encoding="utf-8",
        )

    def _freeze_algorithm_contract(self, workflow: str) -> None:
        (self.test_root / "config" / "algorithm.yaml").write_text(
            yaml.safe_dump(
                {
                    "algorithm": "algorithm",
                    "contract": {
                        "evaluation_workflow": workflow,
                    },
                },
                sort_keys=False,
            ),
            encoding="utf-8",
        )

    def _write_run_receipt(
        self,
        run_index: int,
        status: str,
        *,
        failure_reason: str = "",
    ) -> Path:
        segment_dir = self.test_root / "dataset" / str(run_index)
        (segment_dir / "evaluation").mkdir(parents=True)
        (segment_dir / "receipt.yaml").write_text(
            yaml.safe_dump(
                {
                    "run_index": run_index,
                    "status": status,
                    "failure_reason": failure_reason or None,
                },
                sort_keys=False,
            ),
            encoding="utf-8",
        )
        return segment_dir.resolve()

    def _request(
        self,
        run_index: int,
        workflow: str,
        segment_dir: Path,
        *,
        data_dir: Optional[Path] = None,
        rpe_delta_value: float = 100.0,
        rpe_delta_unit: str = "m",
    ) -> EvaluationRequest:
        selected_data_dir = data_dir or (self.root / f"data-{run_index}")
        selected_data_dir.mkdir(exist_ok=True)
        return EvaluationRequest(
            test_id="test-000",
            algorithm_id="algorithm",
            run_index=run_index,
            dataset_id=f"dataset-{run_index}",
            dataset_type="rk3399",
            segment_id=f"segment-{run_index}",
            workflow=workflow,
            data_dir=selected_data_dir,
            log_dir=segment_dir,
            evaluation_dir=segment_dir / "evaluation",
            rpe_delta_value=rpe_delta_value,
            rpe_delta_unit=rpe_delta_unit,
        )

    def _write_fake_voeval(self) -> Path:
        path = self.root / "fake_voeval.py"
        path.write_text(
            """\
from __future__ import annotations

import json
import sys
from pathlib import Path

mode = sys.argv[1]
data_dir = Path(sys.argv[2])
log_dir = Path(sys.argv[3])
output = Path(sys.argv[sys.argv.index("--output") + 1])
delta_value = float(sys.argv[sys.argv.index("--delta") + 1])
delta_unit = sys.argv[sys.argv.index("--unit") + 1]

if "force-evaluation-failure" in data_dir.name:
    print("intentional evaluator failure")
    raise SystemExit(7)

run_index = int(log_dir.name)
if mode == "sf_vo":
    payload = {
        "mode": mode,
        "rpe_translation_m": {
            "delta_value": delta_value,
            "delta_unit": delta_unit,
            "rmse": 1.0,
            "mean": 2.0,
            "median": 3.0,
            "max": 4.0,
            "min": 0.5,
            "count": 6,
        },
        "segment_count": 3,
    }
else:
    horizontal = (10.0, 20.0, 25.0, 50.0, 55.0)[run_index]
    payload = {
        "mode": mode,
        "vloc_metrics": {
            "trajectory_length_m": 1000.0 + run_index,
            "mean_error_pos_xy": horizontal,
            "mean_error_pos_z": 2.0,
            "mean_error_euler": 3.0,
            "max_error_pos_xy": horizontal + 5.0,
            "max_error_pos_z": 4.0,
            "max_error_euler": 5.0,
        },
    }

output.parent.mkdir(parents=True, exist_ok=True)
output.write_text(json.dumps(payload), encoding="utf-8")
print(f"fake voeval {mode}")
""",
            encoding="utf-8",
        )
        return path

    @staticmethod
    def _yaml(path: Path):
        return yaml.safe_load(path.read_text(encoding="utf-8"))


if __name__ == "__main__":
    unittest.main()
